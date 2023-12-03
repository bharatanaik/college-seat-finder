import openpyxl, django, os
from tqdm import tqdm
from django.db.models import Model

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from main.models import *

class UploadAPI:
    DATABASE:dict[str, Model] = {
        "kcet-data":Cutoff,
        "seat-data":SeatMatrix,
        "jee-orcr":JEEORCR,
        "neet-seat-data":NEETSeatMatrix
    }

    def __init__(self, file:str, type:str) -> None:
        self.wb = openpyxl.load_workbook(file, read_only=True)
        self.model = self.DATABASE[type]
        input("continue if you are okay with data erase.... (ENTER)... or CTRL+C")
        self.model.objects.all().delete()

    def upload(self):
        fields = [i.name for i in self.model._meta.fields]
        fields.remove("id")
        objects = []
        for sheet in self.wb.sheetnames:
            sheet = self.wb[sheet]
            print(f"\n[+] Found {sheet.max_row} in sheer {sheet.title}")
            for row in tqdm(sheet.iter_rows(min_row=2, values_only = True), total=sheet.max_row):
                data = dict(zip(fields, row))
                objects.append(self.model(**data))
        self.model.objects.bulk_create(objects)
        print("\nsaved successfully.. closing the program..")


UploadAPI("excel\JEE ORCR.xlsx", "jee-orcr").upload()