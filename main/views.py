from django.forms.models import BaseModelForm
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.generic import *
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from main.api.kcet import *
from main.api.upload import UploadAPI
from main.api.jee import JEEAPI
from main.api.neet import NEETSeatMatrixAPI
from main.models import Admission



class UploadView:
    @csrf_exempt
    def main(self, request):
        if request.user.is_superuser:
            if request.method == "POST":
                file_type = request.POST.get("file_type")
                file = request.FILES.get("upload_file")
                return JsonResponse(UploadAPI(file, file_type).run())
            return render(request, "upload.html")
        else:
            return render(request, "index.html")

    def delete(self, request, name):
        UploadAPI.DATABASE[name].objects.all().delete()
        return redirect(reverse("upload"))
        
class IndexView(TemplateView):
    template_name = "index.html"

class CutoffAnalyserView(CutoffAPI):
    def main(self, request):
        return render(request, "kcet/cutoff.html", {"types":self._types(),"categories":self._categories()})

    @csrf_exempt
    def get_courses(self, request):
        return JsonResponse(self._courses(**request.POST))

    @csrf_exempt
    def get_data(self, request):
        return JsonResponse(self._cutoff(**request.POST))



class CollegeCutoffView(CutoffAPI):
    def main(self, request):
        return render(request, "kcet/cutoff-college.html", {"types":self._types(), "categories":self._categories})

    @csrf_exempt
    def get_colleges(self, request):
        return JsonResponse(self._colleges(**request.POST))

    @csrf_exempt
    def get_data(self, request):        
        return JsonResponse(self._college_cutoff(**request.POST))


class SeatAnalyserView(SeatAPI):    
    def main(self, request):    
        return render(request, "kcet/seat-finder.html", {"types":self._types(), "categories":self._categories(), "seat_types":self._seat_types()})

    @csrf_exempt
    def get_courses(self, request):
        return JsonResponse(self._courses(**request.POST))

    @csrf_exempt
    def get_districts(self, request):
        return JsonResponse(self._districts(**request.POST))

    @csrf_exempt
    def get_colleges(self, request):
        return JsonResponse(self._colleges(**request.POST))


    @csrf_exempt
    def get_seat_matrix(self, request):
        return JsonResponse(self._matrix(**request.POST))

class JEEORCRView(JEEAPI):
    def main(self, request):
        return render(request, self.get_template(request.GET.get("mode")), self.OPTIONS)

    @csrf_exempt
    def get_institutes(self, request):
        return JsonResponse(self._institutes(**request.POST))

    @csrf_exempt
    def get_courses(self, request):
        return JsonResponse(self._courses(**request.POST))
    
    @csrf_exempt
    def get_result(self, request):
        return JsonResponse(self._result(**request.POST))
            
    @csrf_exempt
    def get_college_result(self, request):
        return JsonResponse(self._courses_via_college(**request.POST))

    @csrf_exempt
    def get_course_result(self, request):
        return JsonResponse(self._colleges_via_course(**request.POST))

class CourseView(object):
    def main(self, request):
        return render(request, "kcet/courses.html", {
            "types":CutoffAPI()._types()
        })
    

class CollegeView:
    def main(self, request, college_name):
        try:
            seats = SeatMatrix.objects.filter(_college_name__icontains = college_name)
            if seats:
                college = seats.first()
            else:
                cutoff = Cutoff.objects.filter(_college_name__icontains = college_name) 
                college = cutoff.first()
                seats = cutoff
            context = {
                "college":{x[1:]:y for x, y in college.__dict__.items() },
                "courses":seats.values_list("_course_name", flat=True).distinct()
            }
            return render(request, "kcet/college.html", context)
        except Exception as e:
            context = {"college":False}
            return render(request, "kcet/college.html", context)
            


class NEETSeatMatrixView(NEETSeatMatrixAPI):
    def main(self, request):
        institute = request.GET.get("institute")
        quota = request.GET.get("quota")
        if institute and quota:
            data = self.get_seat_data(**request.GET)
            print(data)
        return render(request, "neet/seat-matrix.html", self.OPTIONS)
    
    @csrf_exempt
    def get_institutes(self, request):
        return JsonResponse(self._institutes(**request.POST))
    
    @csrf_exempt
    def get_quota(self, request):
        return JsonResponse(self._quota(**request.POST))
    
    @csrf_exempt
    def get_seat_data(self, request):
        return JsonResponse(self._seat_data(**request.POST))
    


def admissions(request):
    context = {}
    if request.method == "POST":
        Admission(course = request.POST["course"], college=request.POST["college"], email=request.POST["email"], whatsapp_no=request.POST["whatsapp_no"]).save()
        context["submit"] = True
    return render(request, "admissions.html", context)



from io import BytesIO

def generate_excel_file():
    workbook = Workbook()
    sheet = workbook.active
    # Add column headers
    sheet['A1'] = 'Course'
    sheet['B1'] = 'College'
    sheet['C1'] = 'Email'
    sheet['D1'] = 'WhatsApp Number'
    admissions = Admission.objects.all()
    row_num = 2 
    for admission in admissions:
        sheet.cell(row=row_num, column=1, value=admission.course)
        sheet.cell(row=row_num, column=2, value=admission.college)
        sheet.cell(row=row_num, column=3, value=admission.email)
        sheet.cell(row=row_num, column=4, value=admission.whatsapp_no)
        row_num += 1
    return workbook

def excel_export_view(request):
    workbook = generate_excel_file()
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=admissions_report.xlsx'
    return response
